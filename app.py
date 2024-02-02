from flask import Flask, render_template, send_file, request, redirect, url_for, session
import os
import pymysql
import pandas as pd
import requests
from io import BytesIO
from werkzeug.utils import secure_filename
import openpyxl

app = Flask(__name__)
app.secret_key = "key"  # Necesario para usar sesiones

# Obtén la ruta del directorio actual del script
script_dir = os.path.dirname(os.path.abspath(__file__))

# -----------------------CONEXION A BASE DE DATOS---------------------------------------


def conectar_base_datos():
    try:
        if os.getenv("ENV") == "PROD":
            # Configurar la conexión a la base de datos (segunda opción)
            db_host = os.getenv("DDBB_HOST")
            db_user = os.getenv("DDBB_USER")
            db_password = os.getenv("DDBB_PASSWORD")

        else:
            # Configurar la conexión a la base de datos
            db_host = "localhost"
            db_user = "root"
            db_password = ""

        # Intentar establecer la conexión
        db_name = "evaluacion_docente"

        connection = pymysql.connect(
            host=db_host, user=db_user, password=db_password, database=db_name
        )
        print("Conexión exitosa a la base de datos")
        return connection

    except pymysql.Error as e:
        print(f"Error al conectar a la base de datos: {e}")
        return None


# -------------------------------FUNCION PARA DESCARGAR EXCEL-------------------------------------------
@app.route("/descargar_informe_final")
def descargar_informe_final():
    # Conectar a la base de datos
    connection = conectar_base_datos()
    if connection is None:
        return "Error de conexión a la base de datos"

    try:
        # Consulta SQL para obtener los datos
        query = "SELECT * FROM informes_finales"  # Cambia el nombre de la tabla

        # Obtener los datos en un DataFrame de pandas
        df = pd.read_sql_query(query, connection)

        # Crear un archivo Excel en memoria (BytesIO)
        excel_buffer = BytesIO()
        df.to_excel(excel_buffer, index=False)
        excel_buffer.seek(0)

        # Enviar el archivo Excel como respuesta al usuario
        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name="Informes Finales.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    finally:
        # Cerrar la conexión a la base de datos
        connection.close()

# -------------------------------FUNCION PARA DESCARGAR EXCEL-------------------------------------------
@app.route("/descargar_informe_final_duplicados")
def descargar_informe_final_duplicados():
    # Conectar a la base de datos
    connection = conectar_base_datos()
    if connection is None:
        return "Error de conexión a la base de datos"

    try:
        # Consulta SQL para obtener los datos
        query = "SELECT * FROM informes_finales_duplicados"  # Cambia el nombre de la tabla

        # Obtener los datos en un DataFrame de pandas
        df = pd.read_sql_query(query, connection)

        # Crear un archivo Excel en memoria (BytesIO)
        excel_buffer = BytesIO()
        df.to_excel(excel_buffer, index=False)
        excel_buffer.seek(0)

        # Enviar el archivo Excel como respuesta al usuario
        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name="Informes Finales Duplicados.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    finally:
        # Cerrar la conexión a la base de datos
        connection.close()

# -----------------------------FUNCIONES CARGA DE ARCHIVOS PLANOS EVALUACIONES-----------------------------------------------


# -----------------------FUNCION CARGA DE DATOS EVALUACION ESTUDIANTES---------------------------------
@app.route("/cargar_datos_eval_estudiantes", methods=["GET", "POST"])
def cargar_datos_eval_estudiantes():
    if request.method == "POST":
        # Obtener el archivo Excel desde el formulario
        archivo_excel = request.files["archivo_excel"]

        try:
            # Leer el archivo Excel
            wb = openpyxl.load_workbook(archivo_excel)
            hoja = wb.active

            # Conectar a la base de datos
            connection = conectar_base_datos()

            if connection:
                # Crear un cursor
                cur = connection.cursor()

                # Truncate table antes del bucle
                truncate_query = "TRUNCATE TABLE e_estud"
                cur.execute(truncate_query)

                # Iterar sobre las filas del archivo Excel e insertar en la base de datos
                for row in hoja.iter_rows(min_row=2, values_only=True):
                    # Verificar y reemplazar celdas vacías con un valor por defecto
                    row = [value if value is not None else " " for value in row]

                    # Construir la consulta INSERT sin la columna autoincremental
                    columnas = """
                        ID_ENCUESTA_QUSUARIO, ID_GRUPO_DOCENTE, FACULTAD, PROGRAMA, GRUPO,
                        DOCUMENTO_DOCENTE, NOMBRE_DOCENTE, CARGO_DOCENTE, ENCUESTA,
                        ID_OPERARIO_U, FECHA_DILIGENCIAMIENTO, PREGUNTA1, PREGUNTA2,
                        PREGUNTA3, PREGUNTA4, PREGUNTA5, PREGUNTA6, PREGUNTA7, PREGUNTA8,
                        PREGUNTA9, PREGUNTA10, PREGUNTA11, PREGUNTA12, PREGUNTA13,
                        PREGUNTA14, PREGUNTA15, PREGUNTA16, PREGUNTA17, PREGUNTA18,
                        PREGUNTA19, PREGUNTA20, PREGUNTA21, PREGUNTA22, PREGUNTA23,
                        PREGUNTA24, PREGUNTA25, PREGUNTA26, PREGUNTA27, PREGUNTA28,
                        PREGUNTA29, PREGUNTA30, PREGUNTA31, PREGUNTA32, PREGUNTA33,
                        PREGUNTA34, PREGUNTA35, PREGUNTA36, PREGUNTA37, PREGUNTA38,
                        PREGUNTA39, PREGUNTA40
                    """
                    marcadores = ", ".join(["%s"] * len(row))

                    query = (
                        f"INSERT INTO e_estud ({columnas}) VALUES ({marcadores})"
                    )

                    # Ejecutar la consulta
                    cur.execute(query, row)

                # Commit y cerrar la conexión
                connection.commit()
                cur.close()
                connection.close()

            # Mensaje de éxito almacenado en la sesión
            session["message"] = "Carga exitosa. Los datos se han subido correctamente."

            # Redirigir a la ruta carga_exitosa
            return redirect(url_for("carga_exitosa_estud"))

        except Exception as e:
            # Mensaje de error almacenado en la sesión
            session["message"] = f"Error durante la carga de datos: {str(e)}"

    # Redirigir en caso de no ser un método POST o en caso de carga exitosa
    return redirect(url_for("e_estud_py_dashboard"))


@app.route("/carga_exitosa_estud")
def carga_exitosa_estud():
    # Obtener el mensaje de la sesión
    message = session.pop("message", None)
    return render_template("carga_exitosa_estud.html", message=message)


@app.route("/e_estud_py_dashboard")
def e_estud_py_dashboard():
    #return redirect("http://localhost/evaluacion_docente/dashboard/e_estud_py.php")
    return redirect("https://application.colmayor.edu.co/evaluacion_docente/dashboard/e_estud_py.php")

    # -----------------------FUNCION CARGA DE DATOS AUTOEVALUACION CON CATEDRA ---------------------------------


@app.route("/cargar_datos_ae_doc_catedra", methods=["GET", "POST"])
def cargar_datos_ae_doc_catedra():
    if request.method == "POST":
        # Obtener el archivo Excel desde el formulario
        archivo_excel = request.files["archivo_excel"]
        print(archivo_excel)

        try:
            # Leer el archivo Excel
            wb = openpyxl.load_workbook(archivo_excel)
            hoja = wb.active

            # Conectar a la base de datos
            connection = conectar_base_datos()

            if connection:
                # Crear un cursor
                cur = connection.cursor()

                # Truncate table antes del bucle
                truncate_query = "TRUNCATE TABLE ae_docente_catedra"
                cur.execute(truncate_query)

                # Iterar sobre las filas del archivo Excel e insertar en la base de datos
                for row in hoja.iter_rows(min_row=2, values_only=True):
                    # Verificar y reemplazar celdas vacías con un valor por defecto
                    row = [value if value is not None else " " for value in row]

                    # Construir la consulta INSERT sin la columna autoincremental
                    columnas = """
                                ID_ENCUESTA_QUSUARIO, ID_DOCENTE, FACULTAD, PROGRAMA, DOCUMENTO_DOCENTE,
                                NOMBRE_DOCENTE, CARGO_DOCENTE, ENCUESTA, FECHA_DILIGENCIAMIENTO, PREGUNTA1,
                                PREGUNTA2, PREGUNTA3, PREGUNTA4, PREGUNTA5, PREGUNTA6, PREGUNTA7, PREGUNTA8,
                                PREGUNTA9, PREGUNTA10, PREGUNTA11, PREGUNTA12, PREGUNTA13, PREGUNTA14,
                                PREGUNTA15, PREGUNTA16, PREGUNTA17, PREGUNTA18, PREGUNTA19, PREGUNTA20,
                                PREGUNTA21, PREGUNTA22, PREGUNTA23, PREGUNTA24, PREGUNTA25, PREGUNTA26,
                                PREGUNTA27, PREGUNTA28, PREGUNTA29, PREGUNTA30, PREGUNTA31
                            """
                    marcadores = ", ".join(["%s"] * len(row))

                    query = (
                        f"INSERT INTO ae_docente_catedra ({columnas}) VALUES ({marcadores})"
                    )

                    # Ejecutar la consulta
                    cur.execute(query, row)

                # Commit y cerrar la conexión
                connection.commit()
                cur.close()
                connection.close()

            # Mensaje de éxito almacenado en la sesión
            session["message"] = "Carga exitosa. Los datos se han subido correctamente."

            # Redirigir a la ruta carga_exitosa
            return redirect(url_for("carga_exitosa_ae_doc_cat"))

        except Exception as e:
            # Mensaje de error almacenado en la sesión
            session["message"] = f"Error durante la carga de datos: {str(e)}"

    # Redirigir en caso de no ser un método POST o en caso de carga exitosa
    return redirect(url_for("ae_catedra_py_dashboard"))


@app.route("/carga_exitosa_ae_doc_cat")
def carga_exitosa_ae_doc_cat():
    # Obtener el mensaje de la sesión
    message = session.pop("message", None)
    return render_template("carga_exitosa_ae_doc_cat.html", message=message)


@app.route("/ae_catedra_py_dashboard")
def ae_catedra_py_dashboard():
    #return redirect("http://localhost/evaluacion_docente/dashboard/ae_doc_cat_py.php")
    return redirect("https://application.colmayor.edu.co/evaluacion_docente/dashboard/ae_doc_cat_py.php")

    # -----------------------FUNCION CARGA DE DATOS AUTOEVALUACION SIN CATEDRA ---------------------------------


@app.route("/cargar_datos_ae_doc_sin_catedra", methods=["GET", "POST"])
def cargar_datos_ae_doc_sin_catedra():
    if request.method == "POST":
        # Obtener el archivo Excel desde el formulario
        archivo_excel = request.files["archivo_excel"]

        try:
            # Leer el archivo Excel
            wb = openpyxl.load_workbook(archivo_excel)
            hoja = wb.active

            # Conectar a la base de datos
            connection = conectar_base_datos()

            if connection:
                # Crear un cursor
                cur = connection.cursor()

                # Truncate table antes del bucle
                truncate_query = "TRUNCATE TABLE ae_docente_sin_catedra"
                cur.execute(truncate_query)

                # Iterar sobre las filas del archivo Excel e insertar en la base de datos
                for row in hoja.iter_rows(min_row=2, values_only=True):
                    # Verificar y reemplazar celdas vacías con un valor por defecto
                    row = [value if value is not None else " " for value in row]

                    # Construir la consulta INSERT sin la columna autoincremental
                    columnas = """
                        ID_ENCUESTA_QUSUARIO, ID_DOCENTE, FACULTAD, PROGRAMA,
                        DOCUMENTO_DOCENTE, NOMBRE_DOCENTE, CARGO_DOCENTE, ENCUESTA,
                        FECHA_DILIGENCIAMIENTO, PREGUNTA1, PREGUNTA2,
                        PREGUNTA3, PREGUNTA4, PREGUNTA5, PREGUNTA6, PREGUNTA7, PREGUNTA8
                    """
                    marcadores = ", ".join(["%s"] * len(row))

                    query = (
                        f"INSERT INTO ae_docente_sin_catedra ({columnas}) VALUES ({marcadores})"
                    )

                    # Ejecutar la consulta
                    cur.execute(query, row)

                # Commit y cerrar la conexión
                connection.commit()
                cur.close()
                connection.close()

            # Mensaje de éxito almacenado en la sesión
            session["message"] = "Carga exitosa. Los datos se han subido correctamente."

            # Redirigir a la ruta carga_exitosa
            return redirect(url_for("carga_exitosa_ae_doc_sin_cat"))

        except Exception as e:
            # Mensaje de error almacenado en la sesión
            session["message"] = f"Error durante la carga de datos: {str(e)}"

    # Redirigir en caso de no ser un método POST o en caso de carga exitosa
    return redirect(url_for("ae_sin_catedra_py_dashboard"))


@app.route("/carga_exitosa_ae_doc_sin_cat")
def carga_exitosa_ae_doc_sin_cat():
    # Obtener el mensaje de la sesión
    message = session.pop("message", None)
    return render_template("carga_exitosa_ae_doc_sin_cat.html", message=message)


@app.route("/ae_sin_catedra_py_dashboard")
def ae_sin_catedra_py_dashboard():
    #return redirect("http://localhost/evaluacion_docente/dashboard/ae_doc_sin_cat_py.php")
    return redirect("https://application.colmayor.edu.co/evaluacion_docente/dashboard/ae_doc_sin_cat_py.php")

    # -----------------------FUNCION CARGA DE DATOS EVALUACION DECANO PLANTA ---------------------------------


@app.route("/cargar_datos_e_dec_planta", methods=["GET", "POST"])
def cargar_datos_e_dec_planta():
    if request.method == "POST":
        # Obtener el archivo Excel desde el formulario
        archivo_excel = request.files["archivo_excel"]

        try:
            # Leer el archivo Excel
            wb = openpyxl.load_workbook(archivo_excel)
            hoja = wb.active

            # Conectar a la base de datos
            connection = conectar_base_datos()

            if connection:
                # Crear un cursor
                cur = connection.cursor()

                # Truncate table antes del bucle
                truncate_query = "TRUNCATE TABLE e_decano_planta"
                cur.execute(truncate_query)

                # Iterar sobre las filas del archivo Excel e insertar en la base de datos
                for row in hoja.iter_rows(min_row=2, values_only=True):
                    # Verificar y reemplazar celdas vacías con un valor por defecto
                    row = [value if value is not None else " " for value in row]

                    # Construir la consulta INSERT sin la columna autoincremental
                    columnas = """
                        ID_ENCUESTA_QUSUARIO, ID_DOCENTE, FACULTAD, PROGRAMA,
                        DOCUMENTO_EVALUADOR, NOMBRE_EVALUADOR, DOCUMENTO_DOCENTE,NOMBRE_DOCENTE,
                        CARGO_DOCENTE, ENCUESTA, FECHA_DILIGENCIAMIENTO, PREGUNTA1, PREGUNTA2,
                        PREGUNTA3, PREGUNTA4, PREGUNTA5, PREGUNTA6, PREGUNTA7, PREGUNTA8,
                        PREGUNTA9, PREGUNTA10, PREGUNTA11, PREGUNTA12, PREGUNTA13,
                        PREGUNTA14, PREGUNTA15, PREGUNTA16, PREGUNTA17, PREGUNTA18,
                        PREGUNTA19
                    """
                    marcadores = ", ".join(["%s"] * len(row))

                    query = (
                        f"INSERT INTO e_decano_planta ({columnas}) VALUES ({marcadores})"
                    )

                    # Ejecutar la consulta
                    cur.execute(query, row)

                # Commit y cerrar la conexión
                connection.commit()
                cur.close()
                connection.close()

            # Mensaje de éxito almacenado en la sesión
            session["message"] = "Carga exitosa. Los datos se han subido correctamente."

            # Redirigir a la ruta carga_exitosa
            return redirect(url_for("carga_exitosa_e_dec_planta"))

        except Exception as e:
            # Mensaje de error almacenado en la sesión
            session["message"] = f"Error durante la carga de datos: {str(e)}"

    # Redirigir en caso de no ser un método POST o en caso de carga exitosa
    return redirect(url_for("e_dec_planta_py_dashboard"))


@app.route("/carga_exitosa_e_dec_planta")
def carga_exitosa_e_dec_planta():
    # Obtener el mensaje de la sesión
    message = session.pop("message", None)
    return render_template("carga_exitosa_e_dec_planta.html", message=message)


@app.route("/e_dec_planta_py_dashboard")
def e_dec_planta_py_dashboard():
    #return redirect("http://localhost/evaluacion_docente/dashboard/e_dec_planta_py.php")
    return redirect("https://application.colmayor.edu.co/evaluacion_docente/dashboard/e_dec_planta_py.php")

    # -----------------------FUNCION CARGA DE DATOS DECANO CATEDRA ---------------------------------


@app.route("/cargar_datos_e_dec_catedra", methods=["GET", "POST"])
def cargar_datos_e_dec_catedra():
    if request.method == "POST":
        # Obtener el archivo Excel desde el formulario
        archivo_excel = request.files["archivo_excel"]

        try:
            # Leer el archivo Excel
            wb = openpyxl.load_workbook(archivo_excel)
            hoja = wb.active

            # Conectar a la base de datos
            connection = conectar_base_datos()

            if connection:
                # Crear un cursor
                cur = connection.cursor()

                # Truncate table antes del bucle
                truncate_query = "TRUNCATE TABLE e_decano_catedra"
                cur.execute(truncate_query)

                # Iterar sobre las filas del archivo Excel e insertar en la base de datos
                for row in hoja.iter_rows(min_row=2, values_only=True):
                    # Verificar y reemplazar celdas vacías con un valor por defecto
                    row = [value if value is not None else " " for value in row]

                    # Construir la consulta INSERT sin la columna autoincremental
                    columnas = """
                        ID_ENCUESTA_QUSUARIO, ID_DOCENTE, FACULTAD, PROGRAMA,
                        DOCUMENTO_EVALUADOR, NOMBRE_EVALUADOR, DOCUMENTO_DOCENTE,NOMBRE_DOCENTE,
                        CARGO_DOCENTE, ENCUESTA, FECHA_DILIGENCIAMIENTO, PREGUNTA1, PREGUNTA2,
                        PREGUNTA3, PREGUNTA4, PREGUNTA5, PREGUNTA6, PREGUNTA7, PREGUNTA8
                    """
                    marcadores = ", ".join(["%s"] * len(row))

                    query = (
                        f"INSERT INTO e_decano_catedra ({columnas}) VALUES ({marcadores})"
                    )

                    # Ejecutar la consulta
                    cur.execute(query, row)

                # Commit y cerrar la conexión
                connection.commit()
                cur.close()
                connection.close()

            # Mensaje de éxito almacenado en la sesión
            session["message"] = "Carga exitosa. Los datos se han subido correctamente."

            # Redirigir a la ruta carga_exitosa
            return redirect(url_for("carga_exitosa_e_dec_catedra"))

        except Exception as e:
            # Mensaje de error almacenado en la sesión
            session["message"] = f"Error durante la carga de datos: {str(e)}"

    # Redirigir en caso de no ser un método POST o en caso de carga exitosa
    return redirect(url_for("e_dec_catedra_py_dashboard"))


@app.route("/carga_exitosa_e_dec_catedra")
def carga_exitosa_e_dec_catedra():
    # Obtener el mensaje de la sesión
    message = session.pop("message", None)
    return render_template("carga_exitosa_e_dec_catedra.html", message=message)


@app.route("/e_dec_catedra_py_dashboard")
def e_dec_catedra_py_dashboard():
    #return redirect("http://localhost/evaluacion_docente/dashboard/e_dec_catedra_py.php")
    return redirect("https://application.colmayor.edu.co/evaluacion_docente/dashboard/e_dec_catedra_py.php")


if __name__ == "__main__":
    app.run(debug=True)
